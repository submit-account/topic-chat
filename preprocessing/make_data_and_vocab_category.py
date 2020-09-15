import os
import json
import openpyxl
import random
from kacteil_nc import KACTEIL_NC
kacteil = KACTEIL_NC(ner=False, parser=False)

def morp_result(sentence):
    # print(sentence)
    morp_result = kacteil.morpheme_analysis(sentence.replace('\n', ' '))
    open_tags = ['NNP', 'NA', 'SN', 'SL']
    result = ''
    for e in morp_result:
        for token in e:
            token = token.split('/')
            word = '/'.join(token[:-1])
            pos = token[-1]
            if pos in open_tags:
                word = '###'+word
            result += word +' '
        result += '<SP> '
    result = result[:-5]

    return result.strip()

def morp_answer_result(sentence):
    # print(sentence)
    morp_result = kacteil.morpheme_analysis(sentence.replace('\n', ' '))
    open_tags = ['NNP', 'NA', 'SN', 'SL']
    result = ''
    for e in morp_result:
        for token in e:
            token = token.split('/')
            word = '/'.join(token[:-1])
            pos = token[-1]
            if pos in open_tags:
                new_word = ''
                for syllable in word:
                    new_word += '#'+syllable + ' '
                word = new_word.strip()
            result += word +' '
        result += '<SP> '
    result = result[:-5]

    return result.strip()

train_list = []
valid_list = []
normal_count = 0
path_dir = './raw_data/data/spoken/spoken'
file_list = os.listdir(path_dir)
cnt = 0
for path in file_list:
    f = open(path_dir+'/'+path, 'r', encoding='utf8')
    json_data = json.load(f)
    if json_data['metadata']['category'] != '구어 > 사적 대화 > 일상 대화':
        continue
    utter_list = json_data['document'][0]['utterance']
    tmp_list = []
    before_speaker = '-1'
    for utter in utter_list:
        if len(tmp_list) == 0:
            tmp_list.append(utter['form'])
            before_speaker = utter['speaker_id']
        else:
            if before_speaker == utter['speaker_id']:
                tmp_list[-1] += ' ' + utter['form']
            else:
                tmp_list.append(utter['form'])
                before_speaker = utter['speaker_id']


    if len(train_list) < 20000:
        for i in range(0, len(tmp_list) - 3):
            try:
                if tmp_list[i].strip() != '' and tmp_list[i+1].strip() != '' and tmp_list[i+2].strip() != '' and tmp_list[i+3].strip() != '':
                    tttt1 = morp_result(tmp_list[i].strip())
                    tttt2 = morp_result(tmp_list[i + 1].strip())
                    tttt3 = morp_result(tmp_list[i + 2].strip())
                    tttt4 = morp_answer_result(tmp_list[i + 3].strip())
                    if len(tttt1.split(' ')) <= 30 and len(tttt2.split(' ')) <= 30 and len(tttt3.split(' ')) <= 30 and len(tttt4.split(' ')) <= 30:
                        if len(tttt1.split(' ')) >= 5 and len(tttt2.split(' ')) >= 5 and len(tttt3.split(' ')) >= 5 and len(tttt4.split(' ')) >= 5:
                            if tttt1.count('#') < 10 and tttt2.count('#') < 10 and tttt3.count('#') < 10 and tttt4.count('#') < 10:
                                train_list.append(tttt1 + '__eou__' + tttt2 + '__eou__' + tttt3 + '__eou__' + tttt4 + '__eou__' + '	--PAD--	일반\n')
            except:
                print('error')
    else:
        if len(valid_list) > 2000:
            break
        for i in range(0, len(tmp_list) - 3):
            try:
                if tmp_list[i].strip() != '' and tmp_list[i + 1].strip() != '' and tmp_list[i + 2].strip() != '' and \
                        tmp_list[i + 3].strip() != '':
                    tttt1 = morp_result(tmp_list[i].strip())
                    tttt2 = morp_result(tmp_list[i + 1].strip())
                    tttt3 = morp_result(tmp_list[i + 2].strip())
                    tttt4 = morp_answer_result(tmp_list[i + 3].strip())
                    if len(tttt1.split(' ')) <= 30 and len(tttt2.split(' ')) <= 30 and len(tttt3.split(' ')) <= 30 and len(tttt4.split(' ')) <= 30:
                        if len(tttt1.split(' ')) >= 5 and len(tttt2.split(' ')) >= 5 and len(tttt3.split(' ')) >= 5 and len(tttt4.split(' ')) >= 5:
                            if tttt1.count('#') < 10 and tttt2.count('#') < 10 and tttt3.count('#') < 10 and tttt4.count('#') < 10:
                                valid_list.append(tttt1 + '__eou__' + tttt2 + '__eou__' + tttt3 + '__eou__' + tttt4 + '__eou__' + '	--PAD--	일반\n')
            except:
                print('error')
    tmp_list = []
    cnt+=1
    f.close()

def get_data_from_worksheet(load_ws, name):
    header = True
    sen_cnt = 0
    valid_cnt = 0
    start_ind = 0
    utter_list = []
    before_speaker = -1
    for row in load_ws.rows:
        if header:
            header = False
            continue
        if sen_cnt + 1 == int(row[7].value):
            if before_speaker == -1:
                before_speaker = row[6].value
                utter_list.append([str(row[1].value), row[6].value])
            elif before_speaker == row[6].value:
                utter_list[-1][0] += ' ' + str(row[1].value)
            else:
                utter_list.append([str(row[1].value), row[6].value])
                before_speaker = row[6].value
            sen_cnt += 1
        else:
            valid_cnt += 1
            if valid_cnt % 10 == 9:
                for i in range(0, len(utter_list) - 3):
                    try:
                        if utter_list[i + 3][1] == 0:
                            if utter_list[i][0].strip() != '' and utter_list[i + 1][0].strip() != '' and utter_list[
                                i + 2][0].strip() != '' and utter_list[i + 3][0].strip() != '':
                                tttt1 = morp_result(utter_list[i][0].strip())
                                tttt2 = morp_result(utter_list[i + 1][0].strip())
                                tttt3 = morp_result(utter_list[i + 2][0].strip())
                                tttt4 = morp_answer_result(utter_list[i + 3][0].strip())
                                if len(tttt1.split(' ')) <= 50 and len(tttt2.split(' ')) <= 50 and len(
                                        tttt3.split(' ')) <= 50 and len(tttt4.split(' ')) <= 50:
                                    if tttt1.count('#') < 10 and tttt2.count('#') < 10 and tttt3.count('#') < 10 and tttt4.count('#') < 10:
                                        valid_list.append(tttt1 + '__eou__' + tttt2 + '__eou__' +tttt3 + '__eou__' +tttt4+ '__eou__' + '	--PAD--	' + name + '\n')
                    except:
                        print('error')
            else:
                for i in range(0, len(utter_list) - 3):
                    try:
                        if utter_list[i+3][1] == 0:
                            if utter_list[i][0].strip() != '' and utter_list[i + 1][0].strip() != '' and utter_list[i + 2][0].strip() != '' and utter_list[i + 3][0].strip() != '':
                                tttt1 = morp_result(utter_list[i][0].strip())
                                tttt2 = morp_result(utter_list[i + 1][0].strip())
                                tttt3 = morp_result(utter_list[i + 2][0].strip())
                                tttt4 = morp_answer_result(utter_list[i + 3][0].strip())
                                if len(tttt1.split(' ')) <= 50 and len(tttt2.split(' ')) <= 50 and len(
                                        tttt3.split(' ')) <= 50 and len(tttt4.split(' ')) <= 50:
                                    if tttt1.count('#') < 10 and tttt2.count('#') < 10 and tttt3.count('#') < 10 and tttt4.count('#') < 10:
                                        train_list.append(tttt1 + '__eou__' + tttt2 + '__eou__' +tttt3 + '__eou__' +tttt4+ '__eou__' + '	--PAD--	' + name + '\n')
                    except:
                        print('error')
            sen_cnt = 1
            utter_list = []
            before_speaker = row[6].value
            utter_list.append([row[1].value, row[6].value])
print('일반 Done')
print(str(len(train_list)) + '\t' + str(len(valid_list)))

load_wb = openpyxl.load_workbook("./raw_data/data/A 음식점(15,726).xlsx", data_only=True)
load_ws = load_wb['15,726']
get_data_from_worksheet(load_ws, '음식점')
print(str(len(train_list)) + '\t' + str(len(valid_list)))
print('Done')

load_wb = openpyxl.load_workbook("./raw_data/data/B 의류(15,826).xlsx", data_only=True)
load_ws = load_wb['Sheet1']
get_data_from_worksheet(load_ws, '의류')
print(str(len(train_list)) + '\t' + str(len(valid_list)))
print('Done')

load_wb = openpyxl.load_workbook("./raw_data/data/C 학원(4,773).xlsx", data_only=True)
load_ws = load_wb['Sheet1']
get_data_from_worksheet(load_ws, '학원')
print(str(len(train_list)) + '\t' + str(len(valid_list)))
print('Done')

load_wb = openpyxl.load_workbook("./raw_data/data/D 소매점(14,949).xlsx", data_only=True)
load_ws = load_wb['Sheet1']
get_data_from_worksheet(load_ws, '소매점')
print(str(len(train_list)) + '\t' + str(len(valid_list)))
print('Done')

load_wb = openpyxl.load_workbook("./raw_data/data/E 생활서비스(11,087).xlsx", data_only=True)
load_ws = load_wb['Sheet1']
get_data_from_worksheet(load_ws, '생활서비스')
print(str(len(train_list)) + '\t' + str(len(valid_list)))
print('Done')

load_wb = openpyxl.load_workbook("./raw_data/data/F 카페(7,859).xlsx", data_only=True)
load_ws = load_wb['Sheet1']
get_data_from_worksheet(load_ws, '카페')
print(str(len(train_list)) + '\t' + str(len(valid_list)))
print('Done')

load_wb = openpyxl.load_workbook("./raw_data/data/G 숙박업(7,113).xlsx", data_only=True)
load_ws = load_wb['Sheet1']
get_data_from_worksheet(load_ws, '숙박업')
print(str(len(train_list)) + '\t' + str(len(valid_list)))
print('Done')

load_wb = openpyxl.load_workbook("./raw_data/data/H 관광여가오락(4,949).xlsx", data_only=True)
load_ws = load_wb['Sheet1']
get_data_from_worksheet(load_ws, '관광여가오락')
print(str(len(train_list)) + '\t' + str(len(valid_list)))
print('Done')

print('Shuffle start')
random.shuffle(train_list)
random.shuffle(valid_list)
print('Shuffle Done')

fout_train = open('./raw_data/category_main_train_data_v3.txt', 'w', encoding='utf8')
fout_valid = open('./raw_data/category_main_valid_data_v3.txt', 'w', encoding='utf8')

print(len(train_list))
print(len(valid_list))
print(train_list[0])
print(valid_list[0])
print(train_list[1000])
print(valid_list[100])
for text in train_list:
    fout_train.write(text)
for text in valid_list:
    fout_valid.write(text)
fout_train.close()
fout_valid.close()